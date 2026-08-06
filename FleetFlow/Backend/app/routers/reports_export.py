"""
reports_export.py
─────────────────
Full reports & export router:

  GET  /reports/fleet-utilization      — Fleet utilization stats
  GET  /reports/fuel-consumption       — Fuel consumption report
  GET  /reports/driver-performance     — Per-driver performance
  GET  /reports/delivery-performance   — Delivery KPIs
  GET  /reports/maintenance            — Already exists; kept for consistency
  GET  /reports/export/pdf/{report}    — Download PDF report
  GET  /reports/export/excel/{report}  — Download Excel report
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from typing import Optional
from collections import Counter, defaultdict
import io

from app.utils.dependencies import get_db, get_current_user
from app.models.user import User
from app.models.vehicle import Vehicle
from app.models.driver import Driver
from app.models.fuel import FuelRecord
from app.models.shipment import Shipment
from app.models.trip import Trip
from app.models.maintenance import MaintenanceRecord

router = APIRouter(prefix="/reports", tags=["Reports & Export"])


# ═══════════════════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def _fmt_inr(val: float) -> str:
    return f"₹{val:,.2f}"

def _fleet_data(db: Session) -> dict:
    vehicles   = db.query(Vehicle).all()
    total      = len(vehicles)
    available  = sum(1 for v in vehicles if v.current_status == "available")
    in_transit = sum(1 for v in vehicles if v.current_status == "in_transit")
    maintenance= sum(1 for v in vehicles if v.current_status == "maintenance")
    util_rate  = round((in_transit / total * 100), 1) if total else 0.0
    types      = Counter(v.vehicle_type for v in vehicles)
    fuel_types = Counter(v.fuel_type for v in vehicles)
    return {
        "total_vehicles": total,
        "available": available,
        "in_transit": in_transit,
        "under_maintenance": maintenance,
        "utilization_rate_pct": util_rate,
        "by_type": dict(types),
        "by_fuel_type": dict(fuel_types),
    }

def _fuel_data(db: Session) -> dict:
    records = db.query(FuelRecord).all()
    total_liters = sum(r.fuel_quantity for r in records)
    total_cost   = sum(r.fuel_cost for r in records)
    avg_cost_per_liter = (total_cost / total_liters) if total_liters else 0.0

    by_vehicle = defaultdict(lambda: {"liters": 0.0, "cost": 0.0, "fills": 0})
    for r in records:
        by_vehicle[r.vehicle_id]["liters"] += r.fuel_quantity
        by_vehicle[r.vehicle_id]["cost"]   += r.fuel_cost
        by_vehicle[r.vehicle_id]["fills"]  += 1

    vehicles = {v.id: v.plate_number for v in db.query(Vehicle).all()}
    vehicle_breakdown = [
        {
            "vehicle_id": vid,
            "plate_number": vehicles.get(vid, f"#{vid}"),
            "total_liters": round(vals["liters"], 2),
            "total_cost": round(vals["cost"], 2),
            "fill_count": vals["fills"],
        }
        for vid, vals in sorted(by_vehicle.items(), key=lambda x: x[1]["liters"], reverse=True)
    ]
    return {
        "total_fuel_liters": round(total_liters, 2),
        "total_fuel_cost": round(total_cost, 2),
        "avg_cost_per_liter": round(avg_cost_per_liter, 2),
        "total_fill_count": len(records),
        "vehicle_breakdown": vehicle_breakdown,
    }

def _driver_data(db: Session) -> dict:
    drivers = db.query(Driver).all()
    total   = len(drivers)
    available = sum(1 for d in drivers if d.is_available)
    avg_safety = round(sum(d.safety_score for d in drivers) / total, 1) if total else 0.0
    avg_rating = round(sum(d.rating for d in drivers) / total, 1) if total else 0.0
    top_driver = max(drivers, key=lambda d: d.completed_trips_count, default=None)
    driver_rows = [
        {
            "driver_id": d.id,
            "name": d.name,
            "trips_completed": d.completed_trips_count,
            "total_km": round(d.total_distance_km, 1),
            "safety_score": d.safety_score,
            "rating": d.rating,
            "status": "Available" if d.is_available else "On Trip",
            "attendance": d.attendance_status,
        }
        for d in sorted(drivers, key=lambda d: d.completed_trips_count, reverse=True)
    ]
    return {
        "total_drivers": total,
        "available_drivers": available,
        "avg_safety_score": avg_safety,
        "avg_rating": avg_rating,
        "top_driver_name": top_driver.name if top_driver else None,
        "driver_rows": driver_rows,
    }

def _delivery_data(db: Session) -> dict:
    shipments = db.query(Shipment).all()
    total     = len(shipments)
    delivered = sum(1 for s in shipments if s.status == "delivered")
    pending   = sum(1 for s in shipments if s.status == "pending")
    in_transit= sum(1 for s in shipments if s.status == "in_transit")
    cancelled = sum(1 for s in shipments if s.status == "cancelled")
    success_rate = round(delivered / total * 100, 1) if total else 0.0

    # Avg delivery time
    delivered_with_time = [
        s for s in shipments if s.status == "delivered" and s.delivered_at
    ]
    avg_hours = 0.0
    if delivered_with_time:
        total_h = sum((s.delivered_at - s.created_at).total_seconds() / 3600 for s in delivered_with_time)
        avg_hours = round(total_h / len(delivered_with_time), 1)

    trips = db.query(Trip).all()
    return {
        "total_shipments": total,
        "delivered": delivered,
        "pending": pending,
        "in_transit": in_transit,
        "cancelled": cancelled,
        "success_rate_pct": success_rate,
        "avg_delivery_time_hours": avg_hours,
        "total_trips": len(trips),
        "completed_trips": sum(1 for t in trips if t.status == "completed"),
    }

def _maintenance_data(db: Session) -> dict:
    records = db.query(MaintenanceRecord).all()
    total      = len(records)
    completed  = sum(1 for r in records if r.status == "completed")
    scheduled  = sum(1 for r in records if r.status == "scheduled")
    in_progress= sum(1 for r in records if r.status == "in_progress")
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    overdue    = sum(
        1 for r in records
        if r.status in ("scheduled", "in_progress") and r.scheduled_date and r.scheduled_date < today
    )
    total_cost = sum(r.cost or 0 for r in records)
    cats = Counter(r.category for r in records if r.category)
    top_cat = cats.most_common(1)[0][0] if cats else None
    return {
        "total_records": total,
        "completed": completed,
        "scheduled": scheduled,
        "in_progress": in_progress,
        "overdue": overdue,
        "total_cost": round(total_cost, 2),
        "top_category": top_cat,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  JSON REPORT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/fleet-utilization")
def fleet_utilization(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return _fleet_data(db)


@router.get("/fuel-consumption")
def fuel_consumption(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return _fuel_data(db)


@router.get("/driver-performance")
def driver_performance(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return _driver_data(db)


@router.get("/delivery-performance")
def delivery_performance(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return _delivery_data(db)


# /reports/maintenance already defined in reports.py — skip to avoid conflict.


# ═══════════════════════════════════════════════════════════════════════════
#  PDF EXPORT
# ═══════════════════════════════════════════════════════════════════════════

VALID_REPORTS = {"fleet-utilization", "fuel-consumption", "driver-performance", "delivery-performance", "maintenance"}


def _build_pdf(report_name: str, data: dict) -> bytes:
    """Build a clean PDF using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.8*cm, rightMargin=1.8*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=18, fontName="Helvetica-Bold",
                                  alignment=TA_CENTER, spaceAfter=6, textColor=colors.HexColor("#1e293b"))
    sub_style   = ParagraphStyle("sub",   fontSize=10, alignment=TA_CENTER,
                                  spaceAfter=16, textColor=colors.HexColor("#64748b"))
    h2_style    = ParagraphStyle("h2",    fontSize=13, fontName="Helvetica-Bold",
                                  spaceBefore=14, spaceAfter=6, textColor=colors.HexColor("#0f172a"))
    normal      = styles["Normal"]

    DARK  = colors.HexColor("#1e293b")
    LIGHT = colors.HexColor("#f8fafc")
    ACCENT= colors.HexColor("#6366f1")

    story = []
    story.append(Paragraph(f"FleetFlow — {report_name.replace('-', ' ').title()} Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}", sub_style))
    story.append(Spacer(1, 0.3*cm))

    def kv_table(rows):
        tdata = [[Paragraph(f"<b>{k}</b>", normal), Paragraph(str(v), normal)] for k, v in rows]
        t = Table(tdata, colWidths=[7*cm, 9*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHT, colors.white]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 10),
            ("PADDING",    (0, 0), (-1, -1), 6),
        ]))
        return t

    def list_table(headers, rows_data):
        tdata = [headers] + rows_data
        col_w = [16*cm / len(headers)] * len(headers)
        t = Table(tdata, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [LIGHT, colors.white]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("PADDING",      (0, 0), (-1, -1), 5),
            ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
        ]))
        return t

    # ── Fleet Utilization ──
    if report_name == "fleet-utilization":
        story.append(Paragraph("Summary", h2_style))
        story.append(kv_table([
            ("Total Vehicles",        data["total_vehicles"]),
            ("Available",             data["available"]),
            ("In Transit",            data["in_transit"]),
            ("Under Maintenance",     data["under_maintenance"]),
            ("Utilization Rate",      f"{data['utilization_rate_pct']}%"),
        ]))
        if data["by_type"]:
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph("Fleet by Type", h2_style))
            story.append(list_table(
                ["Vehicle Type", "Count"],
                [[k, str(v)] for k, v in data["by_type"].items()]
            ))

    # ── Fuel Consumption ──
    elif report_name == "fuel-consumption":
        story.append(Paragraph("Summary", h2_style))
        story.append(kv_table([
            ("Total Fuel Consumed",   f"{data['total_fuel_liters']} L"),
            ("Total Fuel Cost",       _fmt_inr(data["total_fuel_cost"])),
            ("Avg Cost / Litre",      _fmt_inr(data["avg_cost_per_liter"])),
            ("Total Fill-ups",        data["total_fill_count"]),
        ]))
        if data["vehicle_breakdown"]:
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph("Per-Vehicle Breakdown", h2_style))
            story.append(list_table(
                ["Plate", "Litres", "Cost (₹)", "Fill-ups"],
                [[r["plate_number"], str(r["total_liters"]), _fmt_inr(r["total_cost"]), str(r["fill_count"])]
                 for r in data["vehicle_breakdown"]]
            ))

    # ── Driver Performance ──
    elif report_name == "driver-performance":
        story.append(Paragraph("Summary", h2_style))
        story.append(kv_table([
            ("Total Drivers",         data["total_drivers"]),
            ("Available",             data["available_drivers"]),
            ("Avg Safety Score",      f"{data['avg_safety_score']}/100"),
            ("Avg Rating",            f"{data['avg_rating']}/5"),
            ("Top Performer",         data["top_driver_name"] or "N/A"),
        ]))
        if data["driver_rows"]:
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph("Driver Leaderboard", h2_style))
            story.append(list_table(
                ["Driver", "Trips", "KM", "Safety", "Rating", "Status"],
                [[d["name"], str(d["trips_completed"]), str(d["total_km"]),
                  str(d["safety_score"]), str(d["rating"]), d["status"]]
                 for d in data["driver_rows"]]
            ))

    # ── Delivery Performance ──
    elif report_name == "delivery-performance":
        story.append(Paragraph("Summary", h2_style))
        story.append(kv_table([
            ("Total Shipments",       data["total_shipments"]),
            ("Delivered",             data["delivered"]),
            ("In Transit",            data["in_transit"]),
            ("Pending",               data["pending"]),
            ("Cancelled",             data["cancelled"]),
            ("Success Rate",          f"{data['success_rate_pct']}%"),
            ("Avg Delivery Time",     f"{data['avg_delivery_time_hours']} hrs"),
            ("Total Trips",           data["total_trips"]),
            ("Completed Trips",       data["completed_trips"]),
        ]))

    # ── Maintenance ──
    elif report_name == "maintenance":
        story.append(Paragraph("Summary", h2_style))
        story.append(kv_table([
            ("Total Records",         data["total_records"]),
            ("Completed",             data["completed"]),
            ("Scheduled",             data["scheduled"]),
            ("In Progress",           data["in_progress"]),
            ("Overdue",               data["overdue"]),
            ("Total Cost",            _fmt_inr(data["total_cost"])),
            ("Top Category",          data["top_category"] or "N/A"),
        ]))

    doc.build(story)
    return buf.getvalue()


@router.get("/export/pdf/{report_name}")
def export_pdf(
    report_name: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if report_name not in VALID_REPORTS:
        raise HTTPException(400, f"Unknown report. Valid: {', '.join(sorted(VALID_REPORTS))}")
    try:
        from reportlab.lib.pagesizes import A4  # noqa: F401 — import check
    except ImportError:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    data = _get_report_data(report_name, db)
    pdf_bytes = _build_pdf(report_name, data)
    filename = f"fleetflow_{report_name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════

def _build_excel(report_name: str, data: dict) -> bytes:
    """Build an Excel workbook using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_name.replace("-", " ").title()[:31]

    # Styles
    HEADER_FILL = PatternFill("solid", fgColor="1E293B")
    ROW_FILL_A  = PatternFill("solid", fgColor="F8FAFC")
    ROW_FILL_B  = PatternFill("solid", fgColor="FFFFFF")
    ACCENT_FILL = PatternFill("solid", fgColor="6366F1")
    h_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_f = Font(name="Calibri", bold=True, size=14, color="0F172A")
    kv_label= Font(name="Calibri", bold=True, size=10, color="1E293B")
    border  = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def write_title(row, text):
        ws.cell(row, 1, text).font = title_f
        ws.cell(row, 1).alignment = center_align

    def write_header_row(row, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row, ci, col)
            c.font = h_font; c.fill = HEADER_FILL; c.alignment = center_align; c.border = border

    def write_data_row(row, vals, alt=False):
        fill = ROW_FILL_A if alt else ROW_FILL_B
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row, ci, val)
            c.fill = fill; c.border = border
            c.alignment = Alignment(horizontal="center")

    def autofit(max_col):
        for col in range(1, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    row = 1
    ws.cell(row, 1, f"FleetFlow — {report_name.replace('-', ' ').title()} Report").font = title_f
    row += 1
    ws.cell(row, 1, f"Generated: {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}").font = Font(color="64748B", size=9)
    row += 2

    def kv_section(pairs):
        nonlocal row
        for k, v in pairs:
            lc = ws.cell(row, 1, k); lc.font = kv_label; lc.border = border
            vc = ws.cell(row, 2, str(v)); vc.border = border
            row += 1
        row += 1

    def table_section(headers, rows_data):
        nonlocal row
        write_header_row(row, headers)
        row += 1
        for i, r in enumerate(rows_data):
            write_data_row(row, r, alt=bool(i % 2))
            row += 1
        row += 1

    if report_name == "fleet-utilization":
        kv_section([
            ("Total Vehicles", data["total_vehicles"]),
            ("Available",      data["available"]),
            ("In Transit",     data["in_transit"]),
            ("Under Maintenance", data["under_maintenance"]),
            ("Utilization Rate", f"{data['utilization_rate_pct']}%"),
        ])
        if data["by_type"]:
            table_section(["Vehicle Type", "Count"],
                          [[k, v] for k, v in data["by_type"].items()])
        if data["by_fuel_type"]:
            table_section(["Fuel Type", "Count"],
                          [[k, v] for k, v in data["by_fuel_type"].items()])

    elif report_name == "fuel-consumption":
        kv_section([
            ("Total Fuel (L)",    data["total_fuel_liters"]),
            ("Total Cost (₹)",    data["total_fuel_cost"]),
            ("Avg ₹/Litre",       data["avg_cost_per_liter"]),
            ("Total Fill-ups",    data["total_fill_count"]),
        ])
        table_section(
            ["Plate Number", "Total Litres", "Total Cost (₹)", "Fill-ups"],
            [[r["plate_number"], r["total_liters"], r["total_cost"], r["fill_count"]]
             for r in data["vehicle_breakdown"]]
        )

    elif report_name == "driver-performance":
        kv_section([
            ("Total Drivers",    data["total_drivers"]),
            ("Available",        data["available_drivers"]),
            ("Avg Safety Score", data["avg_safety_score"]),
            ("Avg Rating",       data["avg_rating"]),
            ("Top Performer",    data["top_driver_name"] or "N/A"),
        ])
        table_section(
            ["Name", "Trips", "Total KM", "Safety", "Rating", "Status"],
            [[d["name"], d["trips_completed"], d["total_km"], d["safety_score"], d["rating"], d["status"]]
             for d in data["driver_rows"]]
        )

    elif report_name == "delivery-performance":
        kv_section([
            ("Total Shipments",       data["total_shipments"]),
            ("Delivered",             data["delivered"]),
            ("In Transit",            data["in_transit"]),
            ("Pending",               data["pending"]),
            ("Cancelled",             data["cancelled"]),
            ("Success Rate",          f"{data['success_rate_pct']}%"),
            ("Avg Delivery Time (h)", data["avg_delivery_time_hours"]),
            ("Total Trips",           data["total_trips"]),
            ("Completed Trips",       data["completed_trips"]),
        ])

    elif report_name == "maintenance":
        kv_section([
            ("Total Records",   data["total_records"]),
            ("Completed",       data["completed"]),
            ("Scheduled",       data["scheduled"]),
            ("In Progress",     data["in_progress"]),
            ("Overdue",         data["overdue"]),
            ("Total Cost (₹)",  data["total_cost"]),
            ("Top Category",    data["top_category"] or "N/A"),
        ])

    autofit(max_col=8)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.get("/export/excel/{report_name}")
def export_excel(
    report_name: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if report_name not in VALID_REPORTS:
        raise HTTPException(400, f"Unknown report. Valid: {', '.join(sorted(VALID_REPORTS))}")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    data = _get_report_data(report_name, db)
    excel_bytes = _build_excel(report_name, data)
    filename = f"fleetflow_{report_name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ═══════════════════════════════════════════════════════════════════════════
#  HELPER: route report name → data dict
# ═══════════════════════════════════════════════════════════════════════════

def _get_report_data(report_name: str, db: Session) -> dict:
    mapping = {
        "fleet-utilization":    _fleet_data,
        "fuel-consumption":     _fuel_data,
        "driver-performance":   _driver_data,
        "delivery-performance": _delivery_data,
        "maintenance":          _maintenance_data,
    }
    return mapping[report_name](db)
