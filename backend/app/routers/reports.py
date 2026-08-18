from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import func, inspect
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)

from app import models
from app.database import get_db

router = APIRouter()


# =========================================================
# HELPERS
# =========================================================

def grouped_counts(db, model, field):
    rows = db.query(field, func.count(model.id)).group_by(field).all()
    result = {}

    for value, count in rows:
        if value is None:
            key = "Unspecified"
        elif hasattr(value, "value"):
            key = str(value.value).strip()
        else:
            key = str(value).strip()

        key = key.replace("_", " ").title() or "Unspecified"

        old = next((x for x in result if x.lower() == key.lower()), None)
        result[old or key] = result.get(old or key, 0) + count

    return result


def columns_of(model):
    return [c.name for c in inspect(model).columns]


def value_of(obj, field):
    value = getattr(obj, field, "")
    if value is None:
        return ""
    if hasattr(value, "value"):
        return value.value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def title_field(field):
    return field.replace("_", " ").title()


def lookup(db, model, key, value):
    return {
        getattr(row, key): value_of(row, value)
        for row in db.query(model).all()
        if getattr(row, key, None) is not None
    }


def make_lookups(db):
    return {
        "driver": lookup(db, models.Driver, "id", "name"),
        "vehicle": lookup(db, models.Vehicle, "id", "license_plate"),
        "shipment": lookup(db, models.Shipment, "id", "tracking_number"),
        "user": lookup(db, models.User, "id", "email"),
        "trip": {
            row.id: f"TRIP-{row.id:03d}"
            for row in db.query(models.Trip).all()
        },
    }


def display(obj, field, lookups):
    value = value_of(obj, field)

    if field == "driver_id":
        return lookups["driver"].get(value, value)
    if field == "vehicle_id":
        return lookups["vehicle"].get(value, value)
    if field == "shipment_id":
        return lookups["shipment"].get(value, value)
    if field == "user_id":
        return lookups["user"].get(value, value)
    if field == "trip_id":
        return lookups["trip"].get(value, value)

    return value


REPORTS = [
    ("Users", models.User),
    ("Vehicles", models.Vehicle),
    ("Drivers", models.Driver),
    ("Shipments", models.Shipment),
    ("Trips", models.Trip),
    ("Routes", models.Route),
    ("Fuel Logs", models.FuelLog),
    ("Maintenance", models.MaintenanceRecord),
    ("Driver Assignments", models.DriverAssignment),
    ("Driver Attendance", models.DriverAttendance),
    ("Maintenance Alerts", models.MaintenanceAlert),
    ("Notifications", models.Notification),
    ("Audit Logs", models.AuditLog),
]


def detailed_data(db):
    lookups = make_lookups(db)
    result = {}

    for title, model in REPORTS:
        fields = columns_of(model)
        rows = db.query(model).all()

        result[title] = {
            "headers": [title_field(x) for x in fields],
            "rows": [
                [display(row, field, lookups) for field in fields]
                for row in rows
            ],
        }

    return result


def summary_data(db):
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "vehicles": db.query(models.Vehicle).count(),
        "drivers": db.query(models.Driver).count(),
        "shipments": db.query(models.Shipment).count(),
        "trips": db.query(models.Trip).count(),
        "routes": db.query(models.Route).count(),
        "fuel_logs": db.query(models.FuelLog).count(),
        "maintenance": db.query(models.MaintenanceRecord).count(),
        "driver_assignments": db.query(models.DriverAssignment).count(),
        "driver_attendance": db.query(models.DriverAttendance).count(),
        "maintenance_alerts": db.query(models.MaintenanceAlert).count(),
        "notifications": db.query(models.Notification).count(),
        "audit_logs": db.query(models.AuditLog).count(),

        "fuel_quantity": float(
            db.query(func.coalesce(func.sum(
                models.FuelLog.fuel_quantity
            ), 0)).scalar() or 0
        ),
        "fuel_cost": float(
            db.query(func.coalesce(func.sum(
                models.FuelLog.fuel_cost
            ), 0)).scalar() or 0
        ),
        "maintenance_cost": float(
            db.query(func.coalesce(func.sum(
                models.MaintenanceRecord.service_cost
            ), 0)).scalar() or 0
        ),
        "vehicle_capacity": float(
            db.query(func.coalesce(func.sum(
                models.Vehicle.capacity_weight
            ), 0)).scalar() or 0
        ),
        "cargo_weight": float(
            db.query(func.coalesce(func.sum(
                models.Shipment.weight
            ), 0)).scalar() or 0
        ),

        "vehicle_status": grouped_counts(
            db, models.Vehicle, models.Vehicle.status
        ),
        "driver_status": grouped_counts(
            db, models.Driver, models.Driver.status
        ),
        "shipment_status": grouped_counts(
            db, models.Shipment, models.Shipment.current_status
        ),
        "trip_status": grouped_counts(
            db, models.Trip, models.Trip.trip_status
        ),
        "maintenance_status": grouped_counts(
            db, models.MaintenanceRecord,
            models.MaintenanceRecord.maintenance_status
        ),
    }


# =========================================================
# MAIN JSON REPORT
# =========================================================

@router.get("/")
def operations_report(db: Session = Depends(get_db)):
    return summary_data(db)


# =========================================================
# EXCEL
# =========================================================

def style_sheet(sheet):
    fill = PatternFill("solid", fgColor="1F4E78")

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for cells in sheet.columns:
        letter = get_column_letter(cells[0].column)
        length = max(len(str(c.value or "")) for c in cells)
        sheet.column_dimensions[letter].width = min(max(length + 2, 12), 35)

    sheet.freeze_panes = "A2"


def add_summary_sheet(wb, s):
    ws = wb.create_sheet("Fleet Summary", 0)

    ws.append(["FleetFlow Report"])
    ws.append(["Fleet Management & Logistics Platform"])
    ws.append(["Generated", s["generated_at"]])
    ws.append([])

    ws.append(["Fleet Summary"])
    ws.append(["Category", "Total"])

    totals = [
        ("Vehicles", s["vehicles"]),
        ("Drivers", s["drivers"]),
        ("Shipments", s["shipments"]),
        ("Trips", s["trips"]),
        ("Routes", s["routes"]),
        ("Fuel Logs", s["fuel_logs"]),
        ("Maintenance Records", s["maintenance"]),
        ("Driver Assignments", s["driver_assignments"]),
        ("Driver Attendance", s["driver_attendance"]),
        ("Maintenance Alerts", s["maintenance_alerts"]),
        ("Notifications", s["notifications"]),
        ("Audit Logs", s["audit_logs"]),
    ]

    for row in totals:
        ws.append(list(row))

    ws.append([])
    ws.append(["Operational Totals"])
    ws.append(["Metric", "Value"])
    ws.append(["Fuel Quantity", s["fuel_quantity"]])
    ws.append(["Fuel Cost", s["fuel_cost"]])
    ws.append(["Maintenance Cost", s["maintenance_cost"]])
    ws.append(["Vehicle Capacity", s["vehicle_capacity"]])
    ws.append(["Cargo Weight", s["cargo_weight"]])

    for title, values in [
        ("Vehicle Status", s["vehicle_status"]),
        ("Driver Status", s["driver_status"]),
        ("Shipment Status", s["shipment_status"]),
        ("Trip Status", s["trip_status"]),
        ("Maintenance Status", s["maintenance_status"]),
    ]:
        ws.append([])
        ws.append([title])
        ws.append(["Status", "Count"])
        for name, count in values.items():
            ws.append([name, count])

    style_sheet(ws)


@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    s = summary_data(db)
    details = detailed_data(db)

    wb = Workbook()
    wb.remove(wb.active)

    add_summary_sheet(wb, s)

    for title, data in details.items():
        ws = wb.create_sheet(title[:31])
        ws.append(data["headers"])

        for row in data["rows"]:
            ws.append(row)

        style_sheet(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=FleetFlow_Complete_Report.xlsx"
        },
    )


# =========================================================
# PDF
# =========================================================

def pdf_table(data):
    headers = data["headers"]
    rows = data["rows"]

    # PDF tables become too wide if every database column is included.
    # Keep the first 8 columns; Excel contains every column.
    indexes = list(range(min(len(headers), 8)))

    table = [[headers[i] for i in indexes]]

    for row in rows:
        table.append([
            str(row[i]) if row[i] is not None else ""
            for i in indexes
        ])

    wrapped = []
    for r, row in enumerate(table):
        wrapped.append([
            Paragraph(
                value,
                ParagraphStyle(
                    "Cell",
                    fontSize=6 if r else 7,
                    leading=7,
                ),
            )
            for value in row
        ])

    return wrapped


@router.get("/export/pdf")
def export_pdf(db: Session = Depends(get_db)):
    s = summary_data(db)
    details = detailed_data(db)

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=25,
        rightMargin=25,
        topMargin=30,
        bottomMargin=30,
        title="FleetFlow Complete Report",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=22,
    )

    story = [
        Paragraph("FleetFlow Report", title_style),
        Paragraph(
            "Fleet Management & Logistics Platform",
            styles["Normal"],
        ),
        Paragraph(
            f"Generated: {s['generated_at']}",
            styles["Normal"],
        ),
        Spacer(1, 15),
    ]

    summary_rows = [
        ["Metric", "Value"],
        ["Vehicles", s["vehicles"]],
        ["Drivers", s["drivers"]],
        ["Shipments", s["shipments"]],
        ["Trips", s["trips"]],
        ["Routes", s["routes"]],
        ["Fuel Logs", s["fuel_logs"]],
        ["Maintenance Records", s["maintenance"]],
        ["Driver Assignments", s["driver_assignments"]],
        ["Driver Attendance", s["driver_attendance"]],
        ["Maintenance Alerts", s["maintenance_alerts"]],
        ["Notifications", s["notifications"]],
        ["Audit Logs", s["audit_logs"]],
        ["Fuel Quantity", s["fuel_quantity"]],
        ["Fuel Cost", s["fuel_cost"]],
        ["Maintenance Cost", s["maintenance_cost"]],
        ["Vehicle Capacity", s["vehicle_capacity"]],
        ["Cargo Weight", s["cargo_weight"]],
    ]

    summary_table = Table(summary_rows, repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    story += [
        Paragraph("Fleet Summary", styles["Heading2"]),
        summary_table,
        PageBreak(),
    ]

    for title, data in details.items():
        story.append(
            Paragraph(title, styles["Heading2"])
        )

        if data["rows"]:
            table = Table(
                pdf_table(data),
                repeatRows=1,
            )

            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0),
                 colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0),
                 colors.white),
                ("FONTNAME", (0, 0), (-1, 0),
                 "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1),
                 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1),
                 "TOP"),
            ]))

            story.append(table)
        else:
            story.append(
                Paragraph(
                    "No records available.",
                    styles["Normal"],
                )
            )

        story.append(PageBreak())

    doc.build(story)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=FleetFlow_Complete_Report.pdf"
        },
    )